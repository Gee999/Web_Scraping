#!/usr/bin/env python
# coding: utf-8

# # BEST RATED IMDB MOVIES (TOP 250)

# In[42]:


from bs4 import BeautifulSoup
import requests
import openpyxl
import pandas as pd


# In[65]:


#Loading all the information into a CSV file

excel=openpyxl.Workbook()
sheet=excel.active
sheet.title='Top IMDB Movies'
print (excel.sheetnames)
sheet.append(['Movie Name','Year of Release', 'IMDB Rating'])


# In[43]:


URL=("https://www.imdb.com/chart/top/")


# In[44]:


HEADERS=({'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36','Accept-Language':'en-US,en;q=0.5'})


# In[45]:


webpage=requests.get(URL, headers=HEADERS)


# In[46]:


webpage


# In[47]:


soup=BeautifulSoup(webpage.text,"html.parser")
print(soup)


# In[48]:


movies=soup.find('ul',class_='ipc-metadata-list ipc-metadata-list--dividers-between sc-3a353071-0 wTPeg compact-list-view ipc-metadata-list--base')


# In[12]:


print(movies)


# In[49]:


movies=soup.find('ul',class_='ipc-metadata-list ipc-metadata-list--dividers-between sc-3a353071-0 wTPeg compact-list-view ipc-metadata-list--base').find_all('li')


# In[50]:


#To check the total number of values that I should have in the end

print(len(movies))


# In[51]:


#Creating a loop that will iterate each and every value I want returned

for movie in movies:
    name=movie.find('h3',class_='ipc-title__text').text
    print (name)
    break


# In[52]:


#Creating a loop that will iterate each and every value I want returned
#For the first Rank

for movie in movies:
    
    name=movie.find('h3',class_='ipc-title__text').text
    
    year=movie.find('span',class_='sc-14dd939d-6 kHVqMR cli-title-metadata-item').text
        
    rating=movie.find('span',class_='ipc-rating-star ipc-rating-star--base ipc-rating-star--imdb ratingGroup--imdb-rating').text
    
    print (name, year, rating)
    
    break


# In[68]:


#For all the 250 Ranks

for movie in movies:
    
    name=movie.find('h3',class_='ipc-title__text').text
    
    year=movie.find('span',class_='sc-14dd939d-6 kHVqMR cli-title-metadata-item').text
        
    rating=movie.find('span',class_='ipc-rating-star ipc-rating-star--base ipc-rating-star--imdb ratingGroup--imdb-rating').text
    
    print (name, year, rating)
    
    sheet.append([name, year, rating])
    
    excel.save('Top 250 IMDB Movies.xlsx')


# In[66]:


#Loading all the information into a CSV file

excel=openpyxl.Workbook()
sheet=excel.active
sheet.title='Top IMDB Movies'
print (excel.sheetnames)
sheet.append(['Movie Name','Year of Release', 'IMDB Rating'])


# In[67]:





# In[ ]:




